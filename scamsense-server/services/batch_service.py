import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side

from services.message_service import analyze_single_message

def validate_uploaded_file(file_path):

    # ---------------------------------------------
    # Check if file exists
    # ---------------------------------------------
    if not os.path.exists(file_path):

        raise FileNotFoundError(
            "The selected file does not exist."
        )

    # ---------------------------------------------
    # Read uploaded file
    # ---------------------------------------------
    if file_path.lower().endswith(".csv"):

        df = pd.read_csv(file_path)

    elif file_path.lower().endswith((".xlsx", ".xls")):

        df = pd.read_excel(file_path)

    else:

        raise ValueError(
            "Unsupported file format."
        )

    # ---------------------------------------------
    # Validate required column
    # ---------------------------------------------
    if "Text" not in df.columns:

        raise ValueError(
            "Uploaded dataset must contain a column named 'Text'."
        )

    # ---------------------------------------------
    # Validate dataset is not empty
    # ---------------------------------------------
    if df.empty:

        raise ValueError(
            "The uploaded dataset does not contain any messages."
        )

    # ---------------------------------------------
    # Remove empty or blank messages
    # ---------------------------------------------
    df = df.dropna(subset=["Text"])

    df["Text"] = df["Text"].astype(str).str.strip()

    df = df[df["Text"] != ""]

    # ---------------------------------------------
    # Validate remaining messages
    # ---------------------------------------------
    if df.empty:

        raise ValueError(
            "The 'Text' column does not contain any valid messages to analyse."
        )

    return df


def analyze_uploaded_dataset(file_path):

    print("=" * 60)
    print("ScamSense Batch Scam Detection Service")
    print("=" * 60)

    # ---------------------------------------------
    # Validate uploaded dataset
    # ---------------------------------------------
    print("\n[1/4] Validating uploaded dataset...")

    df = validate_uploaded_file(file_path)

    print(f"✓ Dataset loaded successfully.")
    print(f"✓ Total messages found: {len(df)}")

    # Lists for results
    scam_type = []
    confidence = []
    risk_level = []
    summary = []
    why_suspicious = []
    scam_indicators = []
    safety_advice = []
    recommended_actions = []
    prevention_tips = []

    # ---------------------------------------------
    # Analyse every message
    # ---------------------------------------------
    print("\n[2/4] Starting batch analysis...\n")

    total_messages = len(df)

    for index, message in enumerate(df["Text"], start=1):

        print("-" * 60)
        print(f"Processing message {index} of {total_messages}...")
        print("-" * 60)

        ai_result = analyze_single_message(str(message))

        print(f"✓ Prediction: {ai_result['predicted_scam_type']}")
        print(f"✓ Confidence: {ai_result['confidence_score']}")
        print("✓ AI analysis completed.\n")

        scam_type.append(ai_result["predicted_scam_type"])
        confidence.append(ai_result["confidence_score"])
        risk_level.append(ai_result["risk_level"])
        summary.append(ai_result["summary"])
        why_suspicious.append(ai_result["why_suspicious"])

        scam_indicators.append(
        "• " + "\n\n• ".join(ai_result["scam_indicators"])
        )

        safety_advice.append(
            "• " + "\n\n• ".join(ai_result["safety_advice"])
        )

        recommended_actions.append(
            "• " + "\n\n• ".join(ai_result["recommended_actions"])
        )

        prevention_tips.append(
            "• " + "\n\n• ".join(ai_result["prevention_tips"])
        )

    print(f"Completed {index}/{total_messages} messages.\n")

    # ---------------------------------------------
    # Append results
    # ---------------------------------------------
    print("[3/4] Appending analysis results...")

    df["Predicted Scam Type"] = scam_type
    df["Confidence Score"] = confidence
    df["Risk Level"] = risk_level
    df["Summary"] = summary
    df["Why Suspicious"] = why_suspicious
    df["Scam Indicators"] = scam_indicators
    df["Safety Advice"] = safety_advice
    df["Recommended Actions"] = recommended_actions
    df["Prevention Tips"] = prevention_tips

    print("✓ Results appended successfully.")

    print("\n[4/4] Batch analysis completed successfully!")
    print("=" * 60)
    
    output_folder = "outputs"

    os.makedirs(
        output_folder,
        exist_ok=True
    )

    excel_file = os.path.join(
        output_folder,
        "ScamSense_Analysis.xlsx"
    )

    csv_file = os.path.join(
        output_folder,
        "ScamSense_Analysis.csv"
    )

    # ---------------------------------------------
    # Export Excel Report
    # ---------------------------------------------
    df.to_excel(
        excel_file,
        index=False
    )

    # Load workbook
    wb = load_workbook(excel_file)

    ws = wb.active

    # =============================================
    # Header Style
    # =============================================

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="6C3BFF"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(

        left=Side(style="thin"),

        right=Side(style="thin"),

        top=Side(style="thin"),

        bottom=Side(style="thin")

    )

    # =============================================
    # Style Every Cell
    # =============================================

    for row in ws.iter_rows():

        for cell in row:

            cell.alignment = Alignment(

                wrap_text=True,

                vertical="top"

            )

            cell.border = thin_border

    # =============================================
    # Style Header
    # =============================================

    for cell in ws[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # =============================================
    # Better Column Widths
    # =============================================

    column_widths = {

        "A": 45,   # Original Message

        "B": 20,   # Scam Type

        "C": 18,   # Confidence

        "D": 15,   # Risk Level

        "E": 55,   # Summary

        "F": 65,   # Why Suspicious

        "G": 45,   # Scam Indicators

        "H": 45,   # Safety Advice

        "I": 45,   # Recommended Actions

        "J": 45    # Prevention Tips

    }

    for column, width in column_widths.items():

        ws.column_dimensions[column].width = width

    # =============================================
    # Freeze Header Row
    # =============================================

    ws.freeze_panes = "A2"

    # Save Workbook
    wb.save(excel_file)

    # =============================================
    # Export CSV
    # =============================================

    df.to_csv(
        csv_file,
        index=False
    )
    

    return {

    "file_name": os.path.basename(file_path),

    "total_messages": len(df),

    "successful_analyses": len(df),

    "results": df.to_dict(orient="records"),

    "excel_file": excel_file,

    "csv_file": csv_file

}

    

